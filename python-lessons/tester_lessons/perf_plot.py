# -*- coding: utf-8 -*-

"""
-------------------------------------------------
   File Name：     perf_plot
   Description :
   Author :        patrick
   date：          2019/12/12
-------------------------------------------------
   Change Activity: performance data plot
                   2019/12/12:
-------------------------------------------------
"""
__author__ = 'patrick'

import datetime
from collections import namedtuple

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference

MetricSample = namedtuple('MetricsData', 'start_time height num_tx')

def read_metric_file(path="metrics.log"):
    """
    Read and Parse Metric Log
    """
    with open(path, 'r') as raw_metric:
        lines = raw_metric.readlines()
    metrics_data = []
    for line in lines:
        try:
            parsed_data = line.split(" ")
            start_time = " ".join(parsed_data[0:2])
            height = parsed_data[-2:][0].split("=")[1][0]
            num_tx = parsed_data[-1:][0].split("=")[1]
            metrics_sample = MetricSample(start_time=start_time,
                                          height=int(height),
                                          num_tx=int(num_tx))
            metrics_data.append(metrics_sample)
        except IndexError as e:
            pass

    return metrics_data


def metrics_to_excel(metrics, excel_path="metrics.xlsx"):
    wb = Workbook()
    ws = wb.create_sheet("metric", 0)
    ws.cell(1, 1, "duration")
    ws.cell(1, 2, "time_offset")
    ws.cell(1, 3, "tps")
    ws.cell(1, 4, "height")
    ws.cell(1, 5, "num_tx")
    row_num = len(metrics)
    time_offset = 0
    for index in range(row_num):
        ws.cell(index + 2, 4, metrics[index].height)
        ws.cell(index + 2, 5, metrics[index].num_tx)
        if index == 0:
            ws.cell(index + 2, 1, 0)
            ws.cell(index + 2, 2, 0)
            ws.cell(index + 2, 3, 0)
        else:
            start_time = datetime.datetime.strptime(metrics_data[index - 1].start_time,
                                                    "%Y/%m/%d %H:%M:%S.%f")

            end_time = datetime.datetime.strptime(metrics_data[index].start_time,
                                                  "%Y/%m/%d %H:%M:%S.%f")
            duration = (end_time - start_time).seconds
            time_offset = time_offset + duration
            ws.cell(index + 2, 1, duration)
            ws.cell(index + 2, 2, time_offset)
            ws.cell(index + 2, 3, metrics[index].num_tx / duration)
    add_chart(ws, row_num)
    wb.save(excel_path)


def add_chart(ws, max_row):
    lc = LineChart()
    lc.title = "Chain TPS"
    lc.style = 15
    lc.y_axis.title = "TPS"
    lc.x_axis.title = "Time"
    data = Reference(ws, min_col=3, min_row=1,
                     max_col=3, max_row=max_row)
    lc.add_data(data, titles_from_data=True)
    x_timeoffset = Reference(ws,min_col=2,
                                max_col =2,
                                min_row=2,
                                max_row=max_row)
    lc.set_categories(x_timeoffset)
    ws.add_chart(lc,"G13")


if __name__ == '__main__':
    metrics_data = read_metric_file()
    print(metrics_data)
    metrics_to_excel(metrics_data)
